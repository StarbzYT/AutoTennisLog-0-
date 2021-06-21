# Download the helper library from https://www.twilio.com/docs/python/install
from twilio.rest import Client
from auth import *
from twilio.twiml.messaging_response import MessagingResponse
from flask import Flask, request, redirect
from openpyxl import Workbook, load_workbook

class Messages:
    def __init__(self, account_sid, auth_token):
        # Find your Account SID and Auth Token at twilio.com/console
        self.account_sid = account_sid
        self.auth_token = auth_token
    
    # get message history using twilio
    def _message_history(self):
        client = Client(self.account_sid, self.auth_token)
        # do not include this message
        clear = "Sent from your Twilio trial account - Thanks for the message. Configure your number's SMS URL to change this message.Reply HELP for help.Reply STOP to unsubscribe.Msg&Data rates may apply."
        # message history
        messages = client.messages.list(limit=2)
        history = [message.body for message in messages]
        return history

    # format text message for spreadsheet
    def format_text(self):
        message_history = Messages(account_sid, auth_token)._message_history()
        text_message = message_history[-1]  # last index is personal text
        text = text_message
        result = text.split(",")
        opponent = result[0]
        # if the game goes over two sets
        if len(result) > 4:
            score = result[1:4][0]
        # else if the game only went for two sets
        else:
            score = result[1:3][0]
        winner = result[-1]
        return [opponent, score, winner]

    def insert_data(self):
        data = Messages(account_sid, auth_token).format_text()
        work_book = load_workbook(filename = "singles_tournament_2021.xlsx")
        # grab the active worksheet
        ws = work_book.active
        # get max rows
        max_rows = ws.max_row
        # insert opponent
        ws.cell(row=max_rows+1, column=1).value = data[0]
        # insert score
        ws.cell(row=max_rows+1, column=2).value = data[1]
        # insert winner
        ws.cell(row=max_rows+1, column=3).value = data[2]
        # Save the file
        work_book.save("singles_tournament_2021.xlsx")  # will create the file if not already created

insert = Messages(account_sid, auth_token)
insert.insert_data()